# coding:utf-8

import urllib3
import json
import openpyxl

# Create urllib3 PoolManager object
http = urllib3.PoolManager(headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 ('
                                                  'KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'})
# Request a fake token
token_url = "https://deco-my-tree-web.com/api/v1/user/fake-token"
fake_token = http.request('GET', token_url)
access_token = json.loads(fake_token.data.decode('utf-8'))['access_token']
# Set headers
http = urllib3.PoolManager(headers={'Authorization': 'Bearer ' + access_token,
                                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 ('
                                                  'KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'})
# Login
login_url = "https://deco-my-tree-web.com/api/v1/user/login"
email: str = input("Enter your email: ")
password: str = input("Enter your password: ")
login_data = {"email": email, "password": password}
login = http.request('POST', login_url, fields=login_data)
# If request return 404, then login failed
if login.status == 404:
    if json.loads(login.data.decode('utf-8'))['error'] == "toastPleaseCheckYourEmail":
        print("Login failed! Please check your email!")
        exit(1)
    elif json.loads(login.data.decode('utf-8'))['error'] == "toastPleaseCheckYourPassword":
        print("Login failed! Please check your password!")
        exit(1)
hashed_id = json.loads(login.data.decode('utf-8'))['user']['hashed_id']
name = json.loads(login.data.decode('utf-8'))['user']['name']
access_token = json.loads(login.data.decode('utf-8'))['token']['access_token']
print("Login successful! Welcome, " + name + "!")
print("Your tree's hashed_id is " + hashed_id)
# Get tree data
http = urllib3.PoolManager(headers={'Authorization': 'Bearer ' + access_token,
                                    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 ('
                                                  'KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'})
tree_url = "https://deco-my-tree-web.com/api/v1/message/" + hashed_id + "?by_app=true"
tree_data = http.request('GET', tree_url)
message_count = json.loads(tree_data.data.decode('utf-8'))['owner']['message_count']
print("Your tree has " + str(message_count) + " messages!")
trees_data = json.loads(tree_data.data.decode('utf-8'))['trees']
# If multiple trees detected, then put them together
if len(trees_data) > 1:
    print("Multiple trees detected! Putting them together...")
    for i in range(len(trees_data)):
        if i == 0:
            data = trees_data[i]
        else:
            data['messages'] += trees_data[i]['messages']
else:
    data = trees_data[0]
# Write data to Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Tree"
ws['A1'] = "Name"
ws['B1'] = "Message"
for i in range(len(data['messages'])):
    ws['A' + str(i + 2)] = data['messages'][i]['name']
    ws['B' + str(i + 2)] = data['messages'][i]['content']
wb.save("tree.xlsx")
print("Your tree's data has been dumped to tree.xlsx!")
print("Merry Christmas!")
exit(0)
